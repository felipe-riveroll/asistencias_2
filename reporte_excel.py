"""
Módulo para generación de reportes en formato Excel
Utiliza openpyxl para crear archivos Excel con formato profesional
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
from datetime import datetime
from config import TOLERANCIA_RETARDO_MINUTOS, UMBRAL_FALTA_INJUSTIFICADA_MINUTOS


class GeneradorReporteExcel:
    """Clase para generar reportes de asistencia en formato Excel"""

    def __init__(self):
        """Inicializar el generador de reportes Excel"""
        self.workbook = None
        self.worksheet = None

    def crear_reporte_asistencia(
        self, df_reporte, df_resumen, sucursal, periodo_inicio, periodo_fin
    ):
        """
        Crear un reporte completo de asistencia en Excel

        Args:
            df_reporte (pd.DataFrame): DataFrame con datos detallados de asistencia
            df_resumen (pd.DataFrame): DataFrame con resumen del período
            sucursal (str): Nombre de la sucursal
            periodo_inicio (str): Fecha de inicio del período
            periodo_fin (str): Fecha de fin del período

        Returns:
            str: Ruta del archivo Excel generado
        """
        self.workbook = Workbook()

        # Crear hojas de trabajo
        self._crear_hoja_resumen(df_resumen, sucursal, periodo_inicio, periodo_fin)
        self._crear_hoja_detalle(df_reporte)
        self._crear_hoja_estadisticas(df_reporte, df_resumen)

        # Eliminar hoja por defecto
        if "Sheet" in self.workbook.sheetnames:
            self.workbook.remove(self.workbook["Sheet"])

        # Generar nombre de archivo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_archivo = f"reporte_asistencia_{sucursal}_{periodo_inicio}_{periodo_fin}_{timestamp}.xlsx"

        # Guardar archivo
        self.workbook.save(nombre_archivo)

        return nombre_archivo

    def _crear_hoja_resumen(self, df_resumen, sucursal, periodo_inicio, periodo_fin):
        """Crear hoja con resumen ejecutivo"""
        ws = self.workbook.active
        ws.title = "Resumen Ejecutivo"

        # Título principal
        ws["A1"] = f"Reporte de Asistencia - {sucursal}"
        ws["A1"].font = Font(size=16, bold=True)
        ws["A2"] = f"Período: {periodo_inicio} al {periodo_fin}"
        ws["A2"].font = Font(size=12, italic=True)

        # Agregar datos del resumen
        start_row = 5
        for r_idx, row in enumerate(
            dataframe_to_rows(df_resumen, index=False, header=True)
        ):
            for c_idx, value in enumerate(row):
                cell = ws.cell(row=start_row + r_idx, column=c_idx + 1, value=value)

                # Formato para encabezados
                if r_idx == 0:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(
                        start_color="366092", end_color="366092", fill_type="solid"
                    )
                    cell.font = Font(color="FFFFFF", bold=True)
                    cell.alignment = Alignment(horizontal="center")

        self._aplicar_bordes(
            ws, start_row, start_row + len(df_resumen), 1, len(df_resumen.columns)
        )
        self._ajustar_ancho_columnas(ws)

    def _crear_hoja_detalle(self, df_reporte):
        """Crear hoja con datos detallados según formato FRAPPE"""
        ws = self.workbook.create_sheet("Detalle Asistencia")

        # Definir fills de colores según ICG
        fill_azul = PatternFill(
            start_color="3498DB", end_color="3498DB", fill_type="solid"
        )
        fill_amarillo = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )
        fill_rojo = PatternFill(
            start_color="FF0000", end_color="FF0000", fill_type="solid"
        )
        fill_morado = PatternFill(
            start_color="7030A0", end_color="7030A0", fill_type="solid"
        )
        fill_no_laborable = PatternFill(
            start_color="6600CC", end_color="6600CC", fill_type="solid"
        )
        fill_gris = PatternFill(
            start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
        )
        fill_verde_claro = PatternFill(
            start_color="92D050", end_color="92D050", fill_type="solid"
        )
        fill_verde_entrada_nocturno = PatternFill(
            start_color="70AD47", end_color="70AD47", fill_type="solid"
        )

        # Preparar datos con estructura FRAPPE exacta
        df_frappe = self._preparar_datos_frappe(df_reporte)

        # Encabezados FRAPPE
        headers = [
            "ID Empleado",
            "Nombre del empleado",
            "Turno",
            "Fecha",
            "Día",
            "Horas esperadas",
            "Horas totales",
            "Horas Descanso",
            "Checada 1",
            "Checada 2",
            "Checada 3",
            "Checada 4",
            "Checada 5",
            "OBSERVACIONES",
        ]

        # Escribir encabezados
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = fill_azul
            cell.alignment = Alignment(horizontal="center")

        # Procesar datos por empleado
        current_row = 2
        employees = df_frappe["employee"].unique()

        for employee in employees:
            emp_data = df_frappe[df_frappe["employee"] == employee].copy()
            emp_data = emp_data.sort_values("dia")

            # Escribir datos del empleado
            for _, row_data in emp_data.iterrows():
                # Datos básicos
                ws.cell(row=current_row, column=1, value=row_data["employee"])
                ws.cell(row=current_row, column=2, value=row_data["Nombre"])
                ws.cell(
                    row=current_row, column=3, value=row_data.get("turno", "Diurno")
                )

                # Fecha con formato
                fecha_cell = ws.cell(row=current_row, column=4, value=row_data["dia"])
                fecha_cell.number_format = "DD/MM/YYYY"

                # Aplicar color para fechas
                self._aplicar_color_fecha(
                    fecha_cell,
                    row_data["dia"],
                    row_data["dia_semana"],
                    row_data.get("tipo_retardo", ""),
                    row_data.get("tiene_permiso", False),
                    fill_morado,
                    fill_no_laborable,
                    fill_verde_claro,
                )

                ws.cell(row=current_row, column=5, value=row_data["dia_semana"])
                ws.cell(
                    row=current_row, column=6, value=row_data.get("horas_esperadas", "")
                )
                ws.cell(
                    row=current_row,
                    column=7,
                    value=row_data.get("horas_trabajadas", ""),
                )
                ws.cell(
                    row=current_row,
                    column=8,
                    value=row_data.get("horas_descanso", ""),
                )

                # Checadas con formato de color
                checadas = [
                    "checado_1",
                    "checado_2",
                    "checado_3",
                    "checado_4",
                    "checado_5",
                ]
                for idx, checada in enumerate(checadas, 9):
                    checada_value = row_data.get(checada, "")
                    checada_cell = ws.cell(
                        row=current_row, column=idx, value=checada_value
                    )

                    # Aplicar color según hora de checada
                    self._aplicar_color_checada(
                        checada_cell,
                        checada_value,
                        idx == 9,  # idx=9 corresponde a la primera checada (columna I)
                        row_data.get("retardo_perdonado", False),
                        row_data.get("tipo_retardo", ""),
                        row_data.get("hora_entrada_programada", ""),
                        fill_amarillo,
                        fill_rojo,
                        fill_verde_claro,
                        fill_verde_entrada_nocturno,
                    )

                # Observaciones
                observaciones = self._generar_observaciones(row_data)
                ws.cell(row=current_row, column=14, value=observaciones)

                # Aplicar alineación centrada
                for col in range(1, 15):
                    ws.cell(row=current_row, column=col).alignment = Alignment(
                        horizontal="center"
                    )

                current_row += 1

            # Agregar fila de totales por empleado
            self._agregar_fila_totales(ws, current_row, emp_data, fill_gris)
            current_row += 1

        # Aplicar bordes y ajustar columnas
        self._aplicar_bordes(ws, 1, current_row - 1, 1, 14)
        self._ajustar_anchos_frappe(ws)

    def _crear_hoja_estadisticas(self, df_reporte, df_resumen):
        """Crear hoja con estadísticas y KPIs individuales"""
        ws = self.workbook.create_sheet("Estadísticas")

        # Título
        ws["A1"] = "Análisis de KPIs por Empleado"
        ws["A1"].font = Font(size=14, bold=True)
        
        # Información del período
        ws["A2"] = "Indicadores de Rendimiento Individual"
        ws["A2"].font = Font(size=11, italic=True)

        try:
            # Crear analyzer y calcular KPIs usando df_resumen directamente
            analyzer = AsistenciaAnalyzer(dias_laborables_mes=22)
            analyzer.df_data = df_resumen  # Usar df_resumen directamente
            
            # Calcular KPIs
            df_kpis = analyzer.calculate_individual_kpis()
            
            if not df_kpis.empty:
                # Escribir encabezados de KPIs
                start_row = 4
                headers = ['ID', 'Nombre', 'Bradford Factor', 'Tasa Ausentismo (%)', 
                          'Índice Puntualidad', 'Eficiencia Horas (%)', 'SIC']
                
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=start_row, column=col_idx, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")

                # Escribir datos de KPIs
                for row_idx, (_, kpi_row) in enumerate(df_kpis.iterrows(), start_row + 1):
                    ws.cell(row=row_idx, column=1, value=kpi_row['ID'])
                    ws.cell(row=row_idx, column=2, value=kpi_row['Nombre'])
                    ws.cell(row=row_idx, column=3, value=kpi_row['Bradford_Factor'])
                    ws.cell(row=row_idx, column=4, value=kpi_row['Tasa_Ausentismo'])
                    ws.cell(row=row_idx, column=5, value=kpi_row['Indice_Puntualidad'])
                    ws.cell(row=row_idx, column=6, value=kpi_row['Eficiencia_Horas'])
                    ws.cell(row=row_idx, column=7, value=kpi_row['SIC'])
                    
                    # Aplicar formato condicional al SIC
                    sic_cell = ws.cell(row=row_idx, column=7)
                    sic_value = kpi_row['SIC']
                    
                    if sic_value >= 85:
                        sic_cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")  # Verde
                    elif sic_value >= 70:
                        sic_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Amarillo
                    elif sic_value >= 50:
                        sic_cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # Naranja
                    else:
                        sic_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Rojo
                        sic_cell.font = Font(color="FFFFFF")
                
                # Aplicar bordes
                self._aplicar_bordes(ws, start_row, start_row + len(df_kpis), 1, 7)
                
                # Agregar leyenda de interpretación
                legend_start = start_row + len(df_kpis) + 3
                ws[f"A{legend_start}"] = "Interpretación de KPIs:"
                ws[f"A{legend_start}"].font = Font(size=12, bold=True)
                
                interpretations = [
                    "• Bradford Factor: 0-25 (Excelente), 26-50 (Bueno), 51-100 (Regular), >100 (Crítico)",
                    "• Tasa Ausentismo: <5% (Excelente), 5-10% (Aceptable), >10% (Alto)",
                    "• Índice Puntualidad: >95 (Excelente), 85-95 (Bueno), 70-85 (Regular), <70 (Crítico)",
                    "• Eficiencia Horas: >100% (Excelente), 95-100% (Bueno), 85-95% (Regular), <85% (Crítico)",
                    "• SIC: >85 (Excelente), 70-85 (Bueno), 50-70 (Regular), <50 (Crítico)"
                ]
                
                for i, interpretation in enumerate(interpretations):
                    ws[f"A{legend_start + i + 1}"] = interpretation
                    ws[f"A{legend_start + i + 1}"].font = Font(size=10)
                
            else:
                ws["A4"] = "No se pudieron calcular los KPIs - datos insuficientes"
                ws["A4"].font = Font(size=12, color="FF0000")
        
        except Exception as e:
            ws["A4"] = f"Error al generar estadísticas: {str(e)}"
            ws["A4"].font = Font(size=12, color="FF0000")

        self._ajustar_ancho_columnas(ws)

    def _crear_grafico_barras(self, ws, num_categorias, start_row):
        """Crear gráfico de barras para las estadísticas"""
        chart = BarChart()
        chart.title = "Distribución de Estados de Asistencia"
        chart.x_axis.title = "Estados"
        chart.y_axis.title = "Cantidad"

        # Datos para el gráfico
        data = Reference(
            ws, min_col=2, min_row=start_row, max_row=start_row + num_categorias - 1
        )
        categories = Reference(
            ws, min_col=1, min_row=start_row, max_row=start_row + num_categorias - 1
        )

        chart.add_data(data)
        chart.set_categories(categories)

        # Posicionar gráfico
        ws.add_chart(chart, "D3")

    def _aplicar_bordes(self, ws, start_row, end_row, start_col, end_col):
        """Aplicar bordes a un rango de celdas"""
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)

        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                ws.cell(row=row, column=col).border = border

    def _ajustar_ancho_columnas(self, ws):
        """Ajustar automáticamente el ancho de las columnas"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
            ws.column_dimensions[column_letter].width = adjusted_width

    def _preparar_datos_frappe(self, df_reporte):
        """Preparar datos con estructura FRAPPE"""
        df_frappe = df_reporte.copy()

        # Asegurar que tenemos las columnas necesarias
        required_columns = [
            "employee",
            "Nombre",
            "dia",
            "dia_semana",
            "horas_esperadas",
            "horas_trabajadas",
            "horas_descanso",
        ]
        for col in required_columns:
            if col not in df_frappe.columns:
                df_frappe[col] = ""

        # Asegurar que tenemos las 5 checadas
        for i in range(1, 6):
            checada_col = f"checado_{i}"
            if checada_col not in df_frappe.columns:
                df_frappe[checada_col] = ""

        return df_frappe

    def _aplicar_color_fecha(
        self,
        cell,
        fecha,
        dia_semana,
        tipo_retardo,
        tiene_permiso,
        fill_morado,
        fill_no_laborable,
        fill_verde_claro,
    ):
        """Aplicar color a las fechas según el día de la semana, tipo de día y permisos"""
        if tiene_permiso:
            # Días con permiso tienen prioridad y se marcan en verde claro
            cell.fill = fill_verde_claro
        elif tipo_retardo == "Día no Laborable":
            cell.fill = fill_no_laborable
            cell.font = Font(color="FFFFFF")
        elif dia_semana.lower() in ["sábado", "domingo", "sabado"]:
            cell.fill = fill_morado
            cell.font = Font(color="FFFFFF")

    def _aplicar_color_checada(
        self,
        cell,
        checada_value,
        es_entrada,
        retardo_perdonado,
        tipo_retardo,
        hora_entrada_programada,
        fill_amarillo,
        fill_rojo,
        fill_verde,
        fill_verde_entrada_nocturno,
    ):
        """Aplicar color a las checadas según ICG"""
        # Para "Falta Entrada Nocturno" o entrada faltante, aplicar color especial en la primera checada (vacía)
        if es_entrada and (tipo_retardo == "Falta Entrada Nocturno" or 
                          (pd.isna(checada_value) and "Falta registro de entrada" in str(cell.parent.cell(cell.row, 14).value or ""))):
            if pd.isna(checada_value) or checada_value == "" or checada_value == "---":
                cell.fill = fill_verde_entrada_nocturno
                return

        if pd.isna(checada_value) or checada_value == "" or checada_value == "---":
            return

        try:
            # Convertir a tiempo si es string
            if isinstance(checada_value, str) and ":" in checada_value:
                # Extraer solo la parte de tiempo (HH:MM:SS)
                time_str = (
                    checada_value.split()[0] if " " in checada_value else checada_value
                )
                hora = pd.to_datetime(time_str, format="%H:%M:%S").time()
            else:
                return

            if es_entrada:  # Solo para la primera checada (entrada)
                # Si el retardo está perdonado, NO aplicar ningún color (dejar normal)
                if retardo_perdonado:
                    return  # Sin color para retardos perdonados
                
                # Solo aplicar colores si tenemos hora de entrada programada
                if not hora_entrada_programada or hora_entrada_programada == "":
                    return
                
                # Usar el tipo_retardo ya calculado en data_processor en lugar de recalcular
                if tipo_retardo == "A Tiempo":
                    pass  # Sin color (normal)
                elif tipo_retardo == "Retardo":
                    # Retardo menor NO perdonado: amarillo
                    cell.fill = fill_amarillo
                elif tipo_retardo == "Falta Injustificada":
                    # Retardo mayor NO perdonado: rojo
                    cell.fill = fill_rojo
                    cell.font = Font(color="FFFFFF")
        except Exception:
            pass  # Si no se puede parsear, no aplicar color

    def _generar_observaciones(self, row_data):
        """Generar observaciones basadas en los datos de la fila"""
        # Caso especial para días no contratados, para evitar duplicados en observaciones
        if row_data.get("tipo_permiso") == "No Contratado":
            return "No Contratado"

        observaciones = []

        # Verificar observaciones existentes en el DataFrame
        if "observaciones" in row_data and pd.notna(row_data["observaciones"]):
            observaciones.extend(row_data["observaciones"].split("; "))

        # Verificar tipo de retardo
        if "tipo_retardo" in row_data and row_data["tipo_retardo"] not in [
            "A Tiempo",
            "Día no Laborable",
            "",
        ]:
            observaciones.append(row_data["tipo_retardo"])

        # Verificar permisos
        if "tiene_permiso" in row_data and row_data["tiene_permiso"]:
            tipo_permiso = row_data.get("tipo_permiso", "Permiso")
            observaciones.append(f"Permiso: {tipo_permiso}")

        # Verificar salida anticipada
        if "salida_anticipada" in row_data and row_data["salida_anticipada"]:
            observaciones.append("Salida anticipada")

        # Eliminar duplicados y unir
        return "; ".join(list(set(observaciones))) if observaciones else ""

    def _agregar_fila_totales(self, ws, current_row, emp_data, fill_gris):
        """Agregar fila de totales por empleado"""
        # Datos básicos del empleado
        ws.cell(row=current_row, column=1, value=emp_data.iloc[0]["employee"])
        ws.cell(row=current_row, column=2, value=emp_data.iloc[0]["Nombre"])
        ws.cell(row=current_row, column=3, value="Totales")

        # Calcular totales
        dias_trabajados = len(emp_data[emp_data["horas_trabajadas"] != "---"])
        ws.cell(row=current_row, column=4, value=f"{dias_trabajados} días")

        # Total horas esperadas y trabajadas
        try:
            total_esperadas = sum(
                [
                    self._convertir_a_horas(h)
                    for h in emp_data["horas_esperadas"]
                    if h != "---" and h != ""
                ]
            )
            total_trabajadas = sum(
                [
                    self._convertir_a_horas(h)
                    for h in emp_data["horas_trabajadas"]
                    if h != "---" and h != ""
                ]
            )

            ws.cell(
                row=current_row, column=6, value=self._horas_a_string(total_esperadas)
            )
            ws.cell(
                row=current_row, column=7, value=self._horas_a_string(total_trabajadas)
            )

            # Total horas de descanso
            total_descanso = sum(
                [
                    self._convertir_a_horas(h)
                    for h in emp_data["horas_descanso"]
                    if h != "---" and h != ""
                ]
            )
            ws.cell(
                row=current_row, column=8, value=self._horas_a_string(total_descanso)
            )
        except Exception:
            ws.cell(row=current_row, column=6, value="")
            ws.cell(row=current_row, column=7, value="")
            ws.cell(row=current_row, column=8, value="")

        # Aplicar formato gris a toda la fila
        for col in range(1, 15):
            cell = ws.cell(row=current_row, column=col)
            cell.fill = fill_gris
            cell.alignment = Alignment(horizontal="center")

    def _convertir_a_horas(self, valor):
        """Convertir string de horas a decimal"""
        if pd.isna(valor) or valor == "" or valor == "---":
            return 0

        try:
            if isinstance(valor, str) and ":" in valor:
                parts = valor.split(":")
                return int(parts[0]) + int(parts[1]) / 60 + int(parts[2]) / 3600
            return float(valor)
        except Exception:
            return 0

    def _horas_a_string(self, horas_decimal):
        """Convertir horas decimales a formato HH:MM:SS"""
        if horas_decimal == 0:
            return "00:00:00"

        horas = int(horas_decimal)
        minutos = int((horas_decimal - horas) * 60)
        segundos = int(((horas_decimal - horas) * 60 - minutos) * 60)

        return f"{horas:02d}:{minutos:02d}:{segundos:02d}"

    def _ajustar_anchos_frappe(self, ws):
        """Ajustar anchos de columnas según especificaciones FRAPPE"""
        anchos = {
            "A": 12,  # ID Empleado
            "B": 30,  # Nombre
            "C": 15,  # Turno
            "D": 12,  # Fecha
            "E": 12,  # Día
            "F": 15,  # Horas esperadas
            "G": 15,  # Horas totales
            "H": 15,  # Horas Descanso
            "I": 10,  # Checada 1
            "J": 10,  # Checada 2
            "K": 10,  # Checada 3
            "L": 10,  # Checada 4
            "M": 10,  # Checada 5
            "N": 40,  # Observaciones
        }

        for col_letter, width in anchos.items():
            ws.column_dimensions[col_letter].width = width


def parse_time_to_decimal(time_string):
    """
    Convertir string de tiempo HH:MM:SS a decimal
    
    Args:
        time_string (str): Tiempo en formato HH:MM:SS
        
    Returns:
        float: Horas en formato decimal
    """
    if pd.isna(time_string) or time_string == "" or time_string == "---":
        return 0.0
    
    try:
        if isinstance(time_string, str) and ":" in time_string:
            parts = time_string.split(":")
            hours = int(parts[0])
            minutes = int(parts[1]) if len(parts) > 1 else 0
            seconds = int(parts[2]) if len(parts) > 2 else 0
            return hours + minutes / 60 + seconds / 3600
        return float(time_string)
    except (ValueError, IndexError):
        return 0.0


class AsistenciaAnalyzer:
    """
    Clase para analizar datos de asistencia y calcular KPIs individuales
    """
    
    def __init__(self, dias_laborables_mes=22):
        """
        Inicializar el analizador de asistencia
        
        Args:
            dias_laborables_mes (int): Número de días laborables en el mes
        """
        self.dias_laborables_mes = dias_laborables_mes
        self.df_data = None
    
    def load_data(self, excel_file, sheet_name="Resumen Ejecutivo"):
        """
        Cargar datos desde archivo Excel
        
        Args:
            excel_file (str): Ruta del archivo Excel
            sheet_name (str): Nombre de la hoja con los datos
            
        Returns:
            pd.DataFrame: DataFrame con los datos cargados
        """
        try:
            # Leer archivo Excel desde la fila 6 (índice 5)
            df = pd.read_excel(excel_file, sheet_name=sheet_name, header=5)
            
            # Validar que tenemos las columnas necesarias
            required_columns = [
                'employee', 'Nombre', 'total_horas_trabajadas', 'total_horas_esperadas',
                'total_horas_descontadas_permiso', 'total_retardos', 'faltas_del_periodo',
                'faltas_justificadas', 'total_faltas', 'total_salidas_anticipadas'
            ]
            
            # Verificar columnas por índice si los nombres no coinciden
            if len(df.columns) >= 12:
                df.columns = [
                    'employee', 'Nombre', 'total_horas_trabajadas', 'total_horas_esperadas',
                    'total_horas_descontadas_permiso', 'col5', 'col6', 'total_retardos',
                    'faltas_del_periodo', 'faltas_justificadas', 'total_faltas', 
                    'total_salidas_anticipadas'
                ] + [f'col{i}' for i in range(12, len(df.columns))]
            
            # Filtrar filas con datos válidos
            df = df.dropna(subset=['employee', 'Nombre'])
            
            # Convertir tipos de datos
            df['total_retardos'] = pd.to_numeric(df['total_retardos'], errors='coerce').fillna(0)
            df['faltas_del_periodo'] = pd.to_numeric(df['faltas_del_periodo'], errors='coerce').fillna(0)
            df['faltas_justificadas'] = pd.to_numeric(df['faltas_justificadas'], errors='coerce').fillna(0)
            df['total_faltas'] = pd.to_numeric(df['total_faltas'], errors='coerce').fillna(0)
            df['total_salidas_anticipadas'] = pd.to_numeric(df['total_salidas_anticipadas'], errors='coerce').fillna(0)
            
            self.df_data = df
            return df
            
        except Exception as e:
            print(f"Error al cargar datos: {str(e)}")
            return pd.DataFrame()
    
    def calculate_individual_kpis(self):
        """
        Calcular todos los KPIs individuales para cada empleado
        
        Returns:
            pd.DataFrame: DataFrame con KPIs calculados
        """
        if self.df_data is None or self.df_data.empty:
            return pd.DataFrame()
        
        results = []
        
        for _, row in self.df_data.iterrows():
            try:
                # Convertir tiempos a decimal
                horas_trabajadas = parse_time_to_decimal(row['total_horas_trabajadas'])
                horas_esperadas = parse_time_to_decimal(row['total_horas_esperadas'])
                horas_descontadas = parse_time_to_decimal(row['total_horas_descontadas_permiso'])
                
                # Calcular horas esperadas netas
                horas_esperadas_netas = horas_esperadas - horas_descontadas
                
                # 1. Bradford Factor = S² × D
                # 'S' son los episodios de ausencia (ya calculados)
                # 'D' es el total de días de ausencia injustificada
                episodios_ausencia = int(row.get('episodios_ausencia', 0))
                total_dias_ausentes = int(row['faltas_del_periodo'])
                bradford_factor = (episodios_ausencia ** 2) * total_dias_ausentes
                
                # 2. Tasa de Ausentismo (%)
                tasa_ausentismo = (total_dias_ausentes / self.dias_laborables_mes) * 100 if self.dias_laborables_mes > 0 else 0
                
                # 3. Índice de Puntualidad
                retardos = int(row['total_retardos'])
                salidas_anticipadas = int(row['total_salidas_anticipadas'])
                indice_puntualidad = max(0, 100 - ((retardos * 2) + (salidas_anticipadas * 1.5)))
                
                # 4. Eficiencia de Horas (%)
                eficiencia_horas = (horas_trabajadas / horas_esperadas_netas * 100) if horas_esperadas_netas > 0 else 0
                eficiencia_horas = min(eficiencia_horas, 120)  # Cap at 120%
                
                # 5. SIC (Score Integral de Calidad)
                bradford_invertido = 100 - min(bradford_factor, 100)
                sic = (indice_puntualidad * 0.3) + (eficiencia_horas * 0.4) + (bradford_invertido * 0.3)
                
                results.append({
                    'ID': row['employee'],
                    'Nombre': row['Nombre'],
                    'Bradford_Factor': round(bradford_factor, 2),
                    'Tasa_Ausentismo': round(tasa_ausentismo, 2),
                    'Indice_Puntualidad': round(indice_puntualidad, 2),
                    'Eficiencia_Horas': round(eficiencia_horas, 2),
                    'SIC': round(sic, 2)
                })
                
            except Exception as e:
                print(f"Error calculando KPIs para empleado {row.get('employee', 'N/A')}: {str(e)}")
                # Agregar fila con valores por defecto en caso de error
                results.append({
                    'ID': row.get('employee', 'N/A'),
                    'Nombre': row.get('Nombre', 'N/A'),
                    'Bradford_Factor': 0,
                    'Tasa_Ausentismo': 0,
                    'Indice_Puntualidad': 0,
                    'Eficiencia_Horas': 0,
                    'SIC': 0
                })
        
        return pd.DataFrame(results)


def generar_reporte_excel(
    df_reporte, df_resumen, sucursal, periodo_inicio, periodo_fin
):
    """
    Función principal para generar reporte Excel

    Args:
        df_reporte (pd.DataFrame): DataFrame con datos detallados
        df_resumen (pd.DataFrame): DataFrame con resumen del período
        sucursal (str): Nombre de la sucursal
        periodo_inicio (str): Fecha de inicio
        periodo_fin (str): Fecha de fin

    Returns:
        str: Ruta del archivo Excel generado
    """
    generador = GeneradorReporteExcel()
    archivo_excel = generador.crear_reporte_asistencia(
        df_reporte, df_resumen, sucursal, periodo_inicio, periodo_fin
    )

    print(f"✅ Reporte Excel generado: {archivo_excel}")
    return archivo_excel
